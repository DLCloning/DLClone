import codecs
from typing import List

import re

import os
import csv

import sys
import argparse

import math

import openpyxl
from openpyxl.styles import Font, Alignment

# used for find_method
training_map_file=dict()
pretraining_map_file=dict()
prediction_map_file=dict()
targets_map_file=dict()

# used for retrieving start and end for each method
training_map_se=dict()
pretraining_map_se=dict()
prediction_map_se=dict()
targets_map_se=dict()

### START EXCEL UTILITIES

def column_number_to_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def column_string_to_number(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def create_excel_file():
    wb_obj = openpyxl.Workbook()

    sheets = wb_obj.sheetnames

    for sheet in sheets:
        std = wb_obj[sheet]

        wb_obj.remove(std)

    return wb_obj


def make_bold(sheet, cells):
    for cell in cells:
        sheet[cell].font = Font(bold=True)


def fill_cells(sheet, cells, content):
    for x, y in zip(cells, content):
        sheet[x] = y

def wrap_text(sheet, cells):
    for x in cells:
        # print(x)
        sheet[x].alignment = Alignment(wrapText=True)
        # sheet.cell(x).style.alignment.wrap_text = True

        # sheet[x].style.alignment.wrap_text = True

### END EXCEL UTILITIES

dict_read_file=dict()
dict_read_target_file=dict()

def read_file(path, start_line, end_line, start_method, end_method):
    '''
    read the methods from training set and pretraining dataset
    '''
    if path not in dict_read_file.keys():

        file = open(path, 'r')
        lines = file.readlines()
        file.close()

        lines=[l.replace("\n", "").strip() for l in lines]

        dict_read_file[path]=lines

    else:
        lines=dict_read_file[path]




    return lines[(start_line-1):end_line], lines[(start_method):(end_method+1)]

def read_target_file(path, id_method, map_file):
    '''
    read the method from test set (the target) at the <id_method> position and return it
    '''

    if path not in dict_read_target_file.keys():

        file = open(path, 'r')
        lines = file.readlines()
        file.close()

        lines=[l.replace("\n", "").strip() for l in lines]

        dict_read_target_file[path]=lines
    else:
        lines=dict_read_target_file[path]

    for m in map_file:
        # <file>_<method_id>_<start>_<end>
        parts=m.split("_")
        parts=[int(p.replace("\n","")) for p in parts]
        if int(parts[1]) == id_method:
            start_method=int(parts[2])
            end_method=int(parts[3])
            return lines[(start_method):(end_method+1)], start_method, (end_method+1)

    return None, None, None

def extract_num_lines(line):
    '''
    extract the number of duplicate lines
    Found 2 duplicate lines with fingerprint 091694ec5fabc3ae873c8d4448dd0a1d in the following files:
    '''
    temp=line.split("duplicate")[0]
    temp=temp.replace("Found","").strip()

    return int(temp)

def extract_info_clones(line):
    '''
    extract the cloned lines and the file
    Between lines 26 and 27 in /Users/USER/Desktop/test_nicad/NiCad-6.2/files/file_prediction_1.java
    '''
    parts=line.split(" in ")
    start=int((parts[0].replace("Between lines", "").split("and"))[0].strip())
    end=int((parts[0].replace("Between lines", "").split("and"))[1].strip())
    file_name=(parts[1].strip()).split("/")[-1]

    return start, end, file_name




def find_method(file_curr, start_line, end_line, map_file):
    # file is like file_prediction_0.java, file_1.java, ... so we need to take the number id

    # <map_file[<file>]=(<method_id>, <start>, <end>)

    for l in map_file[file_curr]:

        # print("METHOD FROM {} TO {}, CLONE FROM {} TO {}".format(int(parts[2]), int(parts[3]), start_line, end_line))

        if l[1]<=start_line-1 and l[2] >= end_line-1: # the lines of the clone detector start with 1 instead of 0 => -1

            # print(file, int(parts[0]))

            return file_curr, l[0]


    return -1, -1


def return_element_map(file, id_method, map_file):

    try:
        return map_file["{}_{}".format(str(file), str(id_method))]
    except Exception as e:
        return -1, -1

class all_ids:
    def __init__(self, id_folder):
        '''
        This class allows you to retrieve information about the method and masked method IDs
        '''
        self.id_folder=id_folder

    def get_all_ids(self):
        id_folder=self.id_folder
        file = open(os.path.join(id_folder, "test_length.txt"), 'r')
        self.test_length = [f.replace("\n", "") for f in file.readlines()]
        file.close()

        print("read {} lines".format(len(self.test_length)))


        file = open(os.path.join(id_folder, "test_masked.txt"), 'r')
        self.test_masked = [f.replace("\n", "") for f in file.readlines()]
        file.close()

        print("read {} lines".format(len(self.test_masked)))   

        file = open(os.path.join(id_folder, "test_method.txt"), 'r')
        self.test_method = [f.replace("\n", "") for f in file.readlines()]
        file.close()

        print("read {} lines".format(len(self.test_method)))  

        file = open(os.path.join(id_folder, "test_masked_and_method.txt"), 'r')
        self.test_masked_and_method = [f.replace("\n", "") for f in file.readlines()]
        file.close()

        print("read {} lines".format(len(self.test_masked_and_method)))

        file = open(os.path.join(id_folder, "train_method.txt"), 'r')
        self.train_method = [f.replace("\n", "") for f in file.readlines()]
        file.close()

        print("read {} lines".format(len(self.train_method)))

class test_method():
    def __init__(self, test_folder):
        self.test_folder=test_folder
        self.create_list_targets()

    def create_list_targets(self):
        # read test methods

        test_method_dir=self.test_folder

        test_method_file=[f for f in os.listdir(test_method_dir) if ".java" in f]
        test_method_map=[f for f in os.listdir(test_method_dir) if ".txt" in f]

        # print(test_method_file)

        if len(test_method_file)!=1:
            print("ERROR READING TEST METHOD FILE")
            sys.exit(0)

        test_method_file=os.path.join(test_method_dir, test_method_file[0])

        file = open(test_method_file, 'r')
        lines = file.readlines()
        file.close()

        lines=[l.replace("\n", "").strip() for l in lines]

        print("READ {} lines from test method".format(len(lines)))

        test_method_map=os.path.join(test_method_dir, test_method_map[0])

        file = open(test_method_map, 'r')
        lines_map = file.readlines()
        file.close()

        lines_map=[l.replace("\n", "").strip() for l in lines_map]

        result=list()
        for l in lines_map:
            parts=l.split("_")
            curr_id=parts[1]
            lines_curr=lines[int(parts[2]):int(parts[3])]
            result.append(lines_curr)

        self.test_methods=result


def find_clones(args):

    clone_file=args.simian
    file_dir=args.file_dir
    pretraining_dir=args.pretraining_dir
    predictions_dir=args.target
    target_dir=args.target
    test_method_dir=args.test_method
    id_folder=args.id_folder
    excel_file=args.output_excel
    min_len=args.min
    max_len=args.max

    all_ids_instance=all_ids(id_folder)
    all_ids_instance.get_all_ids()

    # READ SIMIAN OUTPUT
    file = open(clone_file, 'r')
    lines = file.readlines()
    file.close()

    print("READ {} lines from simian file".format(len(lines)))

    # remove the first lines since they are not related to the found clones

    start_line=0
    for i, l in enumerate(lines):
        if l.startswith("Found"):
            start_line=i
            break

    lines=lines[start_line:]

    num_clones=0

    dict_lines=dict()

    dict_clones=dict()
    id_clone=-1

    dict_clones_numlines=dict()

    separator="|||"

    for line in lines:

        if line.startswith("Found"):
            num_lines=extract_num_lines(line)
            id_clone+=1
            dict_clones_numlines[id_clone]=num_lines


            dict_clones[id_clone]=list()
            if num_lines not in dict_lines.keys():
                dict_lines[num_lines]=1
            else:
                dict_lines[num_lines]+=1

        elif line.startswith("Processed"):
            break

        else:
            start, end, name=extract_info_clones(line)
            dict_clones[id_clone].append("{}{}{}{}{}".format(start, separator, end, separator, name))


    print("FOUND {} DIFFERENT CLONES".format(len(list(dict_clones.keys()))))

    write_excel_clones(dict_clones, dict_clones_numlines, separator, file_dir, pretraining_dir, predictions_dir, target_dir, test_method_dir, excel_file, min_len, max_len, all_ids_instance)


def return_test_method(ids_class, test_method_class, curr_method):
    '''
    this function returns the test method containing the block we want to predict
    '''
    test_ids=ids_class.test_masked
    # if curr_method=5, it means that the masked method we're looking for
    # is the 6th one in the list of masked method we want to predict
    # this corrisponds to the 6th masked method id in the list
    curr_test_id=test_ids[curr_method] 

    # we want to get the method associated to the masked method
    masked_and_method=ids_class.test_masked_and_method
    method_curr=-1
    for m in masked_and_method:
        parts=m.split("|||")
        if parts[0]==curr_test_id:
            method_curr=parts[1]
            break


    methods_ids=ids_class.test_method
    
    id_method=-1

    for i, m in enumerate(methods_ids):
        if m == method_curr:
            id_method=i
            break

    # test_method_curr is the entire test method that contains the block
    # we want to predict
    test_method_curr=test_method_class.test_methods[id_method]
    return test_method_curr


def return_method(file_dir, pretraining_dir, prediction_dir, file_cloning, start_cloning, end_cloning):

    curr_map=training_map_file
    curr_map_se=training_map_se

    id_="train"

    if "target" in file_cloning:
        curr_map=prediction_map_file
        curr_map_se=prediction_map_se
        id_="prediction"

    if "pretrain" in file_cloning:
        curr_map=pretraining_map_file
        curr_map_se=pretraining_map_se
        id_="pretrain"

    id_file_cloning=int(file_cloning.replace(".java", "").split("_")[-1])

    # return the file (e.g., 1) and the method (e.g., 12)
    # meaning that it is the method in file_1.txt number 12, starting from 0 (13th line)
    curr_file, curr_method=find_method(id_file_cloning, start_cloning, end_cloning, curr_map)

    if curr_file==-1:
        print("NOT FOUND FILE {} {} {}".format(id_file_cloning, start_cloning, end_cloning))
        return -1, -1, -1, -1, -1

    start_method, end_method= return_element_map(curr_file, curr_method, curr_map_se)

    if start_method == -1:
        print("NOT FOUND METHOD {} {}".format(curr_file, curr_method))
        return -1, -1, -1, -1, -1



    # we saved all the lines in a dict where the key is the path of the file
    key_file=os.path.join(file_dir, file_cloning)
    if id_ == "prediction":
        key_file=os.path.join(prediction_dir, file_cloning)
    elif id_ == "pretrain":
        key_file=os.path.join(pretraining_dir, file_cloning)

    lines, lines_all=read_file(key_file, start_cloning, end_cloning, start_method, end_method)

    return curr_file, curr_method, lines, lines_all, id_

def get_method_masked_method_id(id_dataset, ids_class, curr_method):
    '''
    return the method id. If it is a prediction also the masked method id
    we don't have this information for pretraining methods
    '''
    if id_dataset=="train":
        return ids_class.train_method[curr_method], -1

    elif id_dataset == "pretrain":
        return -1, -1

    else: # prediction
        element=ids_class.test_masked_and_method[curr_method]
        parts=element.split("|||")
        return parts[1], parts[0]

def write_excel_clones(dict_clones, dict_clones_numlines, separator, file_dir, pretraining_dir, prediction_dir, target_dir, test_method_dir, excel_file, min_len, max_len, ids_class):

    wb=create_excel_file()
    wb.create_sheet('Clones')
    sheet = wb["Clones"]

    # READ MAPPING FILE
    file = open(os.path.join(file_dir, "file_map.txt"), 'r')
    training_map = file.readlines()
    file.close()

    file = open(os.path.join(pretraining_dir, "pretraining_file_map.txt"), 'r')
    pretraining_map = file.readlines()
    file.close()

    file = open(os.path.join(prediction_dir, "file_target_map.txt"), 'r')
    prediction_map = file.readlines()
    file.close()

    file = open(os.path.join(target_dir, "file_target_map.txt"), 'r')
    targets_map = file.readlines()
    file.close()

    # preparing maps for find_method and start_end

    for l in prediction_map:
        l_curr=l.replace("\n", "")
        parts=l.split("_")
        parts=[int(p) for p in parts]
        if parts[0] not in prediction_map_file:
            prediction_map_file[parts[0]]=list()

        prediction_map_file[parts[0]].append(parts[1:])
        prediction_map_se["{}_{}".format(str(parts[0]), str(parts[1]))]=parts[2:]


    for l in training_map:
        l_curr=l.replace("\n", "")
        parts=l.split("_")
        parts=[int(p) for p in parts]
        if parts[0] not in training_map_file:
            training_map_file[parts[0]]=list()

        training_map_file[parts[0]].append(parts[1:])
        training_map_se["{}_{}".format(str(parts[0]), str(parts[1]))]=parts[2:]

    for l in pretraining_map:
        l_curr=l.replace("\n", "")
        parts=l.split("_")
        parts=[int(p) for p in parts]
        if parts[0] not in pretraining_map_file:
            pretraining_map_file[parts[0]]=list()

        pretraining_map_file[parts[0]].append(parts[1:])
        pretraining_map_se["{}_{}".format(str(parts[0]), str(parts[1]))]=parts[2:]

    for l in targets_map:
        l_curr=l.replace("\n", "")
        parts=l.split("_")
        parts=[int(p) for p in parts]
        if parts[0] not in targets_map_file:
            targets_map_file[parts[0]]=list()

        targets_map_file[parts[0]].append(parts[1:])
        targets_map_se["{}_{}".format(str(parts[0]), str(parts[1]))]=parts[2:]


    print("START PROCESSING")


    count=0

    id_file=0 # we will save an excel file every 500k lines increasing this value

    curr_row_excel=1

    fill_cells(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "k1", "L1"],
        ["Clone ID", "Internal ID", "Num Lines", "File", "Start Line", "End Line", "Cloned Part", 
        "Code checked for duplicates", "Correct prediction", "Test method", "Method method id", "Masked method id"])

    make_bold(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "k1", "L1"])

    sheet.column_dimensions["G"].width = 50
    sheet.column_dimensions["H"].width = 50
    sheet.column_dimensions["I"].width = 50
    sheet.column_dimensions["J"].width = 50

    # read prediction file

    prediction_file=[f for f in os.listdir(prediction_dir) if ".java" in f]

    if len(prediction_file)!=1:
        print("ERROR READING PREDICTION FILE")
        sys.exit(0)

    prediction_file=os.path.join(prediction_dir, prediction_file[0])

    file = open(prediction_file, 'r')
    lines = file.readlines()
    file.close()

    lines=[l.replace("\n", "").strip() for l in lines]

    dict_read_file[prediction_file]=lines

    test_method_class=test_method(test_method_dir)

    # iterate for all the clones

    for i, k in enumerate(dict_clones.keys()):
        # print("Processing clone {}".format(k))

        if i%50000==0:
            print("processed {} clones out of {}".format(i, len(dict_clones.keys())))

        if dict_clones_numlines[k]>max_len or dict_clones_numlines[k]<min_len:
            continue

        files=list()
        for el in dict_clones[k]:
            parts=el.split(separator)
            files.append(parts[2])

        files_prediction=[f for f in files if "target" in f]

        # there is at least a prediction file in the set and one training file
        # this means that the clone involves both prediction and training files
        if len(files_prediction) > 0 and len(files_prediction) < len(files):
            # print(list(dict.fromkeys(files)))

            count+=1
            # if count>=7:
            #     break

            # each el is like 157605|||157609|||file_1.java (start of the clone, end of the clone, file where the clone is)
            # we have one el for each element in the clone cluster (>=2 elements)
            for el in dict_clones[k]:
                curr_row_excel+=1
                parts=el.split(separator)


                fill_cells(sheet, ["A{}".format(curr_row_excel), "C{}".format(curr_row_excel)],[k, dict_clones_numlines[k]])

                # print("CLONE FROM LINE {} TO LINE {} FOR FILE {}".format(parts[0], parts[1], parts[2]))

                file_cloning=parts[2] # file in which we found the clone
                start_cloning=int(parts[0]) # start line for the clone
                end_cloning=int(parts[1]) # end line for the clone

                # returns
                # curr_file: (e.g., 1 involves file_1.java, 0 involves file_prediction_0.java)
                # curr_method (e.g, 12): the 13th method in that file
                # lines: the lines involved in the cloning
                # lines_all: all the lines where there is the cloning 
                # (it can be the whole method for training or the block for prediction)
                # id_dataset: "train" or "prediction" or "pretrain"
                curr_file, curr_method, lines, lines_all, id_dataset=return_method(file_dir, pretraining_dir, prediction_dir, 
                    file_cloning, start_cloning, end_cloning)

                if curr_method==-1: # problem in retrieving method
                    continue

                fill_cells(sheet, ["B{}".format(curr_row_excel)],[curr_method])

                method_id, masked_method_id=get_method_masked_method_id(id_dataset, ids_class, curr_method)

                fill_cells(sheet, ["K{}".format(curr_row_excel)],[method_id])
                
                # adding masked method id
                if id_dataset=="prediction":
                    fill_cells(sheet, ["L{}".format(curr_row_excel)],[masked_method_id])

                # adding test method that contains the block we want to predict
                if id_dataset=="prediction":

                    # return the test method that contains the block we want to predict
                    test_method_curr = return_test_method(ids_class, test_method_class, curr_method)

                    fill_cells(sheet, ["J{}".format(curr_row_excel)],
                    ["\n".join(test_method_curr)])
                    # wrap_text(sheet, ["J{}".format(curr_row_excel)])


                fill_cells(sheet, ["D{}".format(curr_row_excel), "E{}".format(curr_row_excel),
                    "F{}".format(curr_row_excel)], [file_cloning, start_cloning, end_cloning]) 
                fill_cells(sheet, ["G{}".format(curr_row_excel)], ["\n".join(lines)])
                fill_cells(sheet, ["H{}".format(curr_row_excel)], ["\n".join(lines_all)])

                # wrap_text(sheet, ["G{}".format(curr_row_excel)])
                # wrap_text(sheet, ["H{}".format(curr_row_excel)])


                if id_dataset == "prediction":
                    target_lines, start_target, end_target=read_target_file(os.path.join(target_dir, "file_target_0.java"), 
                        int(curr_method), targets_map)

                    fill_cells(sheet, ["I{}".format(curr_row_excel)],
                    ["\n".join(target_lines)])

                    # wrap_text(sheet, ["I{}".format(curr_row_excel)])

        if curr_row_excel>200000:
            wb.save(excel_file.replace(".xlsx", "_{}.xlsx".format(id_file)))
            wb=create_excel_file()
            wb.create_sheet('Clones')
            sheet = wb["Clones"]
            id_file+=1
            curr_row_excel=1

            fill_cells(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "k1", "L1"],
                ["Clone ID", "Internal ID", "Num Lines", "File", "Start Line", "End Line", "Cloned Part", 
                "Code checked for duplicates", "Correct prediction", "Test method", "Method method id", "Masked method id"])

            make_bold(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "k1", "L1"])

            sheet.column_dimensions["G"].width = 50
            sheet.column_dimensions["H"].width = 50
            sheet.column_dimensions["I"].width = 50
            sheet.column_dimensions["J"].width = 50





    print("FOUND {} DIFFERENT CLONES WITH PREDICTIONS".format(count))

    wb.save(excel_file.replace(".xlsx", "_{}.xlsx".format(id_file)))


if __name__ == "__main__":


    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--simian",
        type=str,
        help="output of simian"
    )
    parser.add_argument(
        "--file_dir",
        type=str,
        help="directory with all the training files passed to simian"
    )

    parser.add_argument(
        "--pretraining_dir",
        type=str,
        help="directory with all the pretraining files passed to simian"
    )

    parser.add_argument(
        "--predictions",
        type=str,
        help="directory with all the predictions"
    )

    parser.add_argument(
        "--target",
        type=str,
        help="target file"
    )

    parser.add_argument(
        "--test_method",
        type=str,
        help="path to the test methods (each test method is the entire method, not only the prediction)"
    )

    parser.add_argument(
        "--id_folder",
        type=str,
        help="folder that contains all the file to map to the method and masked method IDs"
    )
    
    parser.add_argument(
        "--output_excel",
        type=str,
        default="test.xlsx",
        help="output file name for excel"
    )

    parser.add_argument(
        "--min",
        type=int,
        default=5,
        help="minimum number of lines for cloning (<=)"
    )

    parser.add_argument(
        "--max",
        type=int,
        default=999999,
        help="maximum number of lines for cloning (<=)"
    )

    args = parser.parse_args()

    find_clones(args)

