import os

import sys
import argparse

import math

import openpyxl
from openpyxl.styles import Font, Alignment

from utilities_summary2.excel_API import *
from utilities_summary2.prediction_cluster_class import Prediction, Cluster, Oracle
from utilities_summary2.prediction_utilities import *

# folders you want to process
# folders=["test"]
folders=["0302", "0403", "0605", "0807"]
# folders=["0302", "0605", "0807"]
# folders=["0605_FAKE", "0605"]
# folders=["0302", "0605", "0807"]


def return_cells(from_column, end_column, row):
    cells=list()
    for i in range(from_column, end_column+1):
        cell="{}{}".format(column_number_to_string(i), row)
        cells.append(cell)

    return cells

def write_excels(output_folder, columns, dict_result):
    

    for k in dict_result.keys():
        print("writing {}.xlsx".format(k))

        wb=create_excel_file()
        wb.create_sheet('Summary')
        sheet = wb["Summary"]

        fill_cells(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1", "M1", "N1", "O1", "P1", "Q1", "R1", "S1", "T1", "U1", "V1", "W1"],
        columns)

        make_bold(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1", "M1", "N1", "O1", "P1", "Q1", "R1", "S1", "T1", "U1", "V1", "W1"])

        curr_row=2

        for row in dict_result[k]:
            cells=return_cells(1, 23, curr_row)
            print(cells)
            fill_cells(sheet, cells, row)
            curr_row+=1


        wb.save(os.path.join(output_folder, "{}.xlsx".format(k)))

def write_list(filename, list_elements):
    with open(filename, 'w+') as file:
        for el in list_elements:
            file.write('{}\n'.format(el))

def filter_classes(clusters, predictions, clusters_oracle, oracles, remove_unknown, minimum_length):
    '''
    filter all object. If remove_unknown is True, it removes predictions with an unknown token
    if minimum length = k, it keeps only cluster with a length >=k
    '''

    id_predictions=dict()
    id_clusters=dict()
    id_clusters_oracle=dict()

    print("LEN CLUSTERS: {}".format(len(clusters.keys())))

    if remove_unknown == False:
        for key in predictions.keys():
            id_predictions[key]=1
    else:
        for key in predictions.keys():
            if predictions[key].unknown_token == False:
                id_predictions[key]=1
     
    for key in clusters.keys():
        if clusters[key].num_lines>=minimum_length:
            id_clusters[key]=1

    for key in clusters_oracle.keys():
        if clusters_oracle[key].num_lines>=minimum_length:
            id_clusters_oracle[key]=1


    print("LEN ID CLUSTERS: {}".format(len(id_clusters.keys())))

    new_predictions=dict()
    new_clusters=dict()
    new_oracles=dict()
    new_clusters_oracle=dict()

    for id_ in id_predictions.keys():
        p_curr=predictions[id_].clone()
        p_curr.clusters=list()
        for cluster in predictions[id_].clusters:
            if cluster in id_clusters.keys():
                p_curr.clusters.append(cluster)

        new_predictions[id_]=p_curr

    for id_ in id_predictions.keys():
        o_curr=oracles[id_].clone()
        o_curr.clusters=list()
        for cluster in oracles[id_].clusters:
            if cluster in id_clusters_oracle.keys():
                o_curr.clusters.append(cluster)
        
        new_oracles[id_]=o_curr

    # print("PRINT CLUSTERS")
    for id_ in id_clusters.keys():
        # print(id_)
        # print(clusters[id_].pretraining)
        # print(clusters[id_].training)
        # print(clusters[id_].predictions)
        c_curr=clusters[id_].clone()
        c_curr.predictions=list()
        for id_curr in clusters[id_].predictions:
            if id_curr in id_predictions.keys():
                c_curr.predictions.append(id_curr)
        # print(len(c_curr.predictions))
        if len(c_curr.predictions)>0: # if at least one prediction is in the cluster (we may remove the ones with the unknown token)
            new_clusters[id_]=c_curr

    for id_ in id_clusters_oracle.keys():
        c_curr=clusters_oracle[id_].clone()
        c_curr.predictions=list()
        for id_curr in clusters_oracle[id_].predictions:
            if id_curr in id_predictions.keys():
                c_curr.predictions.append(id_curr)

        if len(c_curr.predictions)>0: # if at least one prediction is in the cluster (we may remove the ones with the unknown token)
            new_clusters_oracle[id_]=c_curr

    return new_clusters, new_predictions, new_clusters_oracle, new_oracles

def merge_type1_type2(result_type1, result_type2):
    '''
    it merges the result from type1 clones and type2  clones
    '''
    result_merged=result_type1.copy()
    for k in result_type2[3:]:
        result_merged.append(k)
    return result_merged

def create_all_excels(clusters_type1, clusters_type2, predictions_type1, predictions_type2,
            clusters_type1_oracle, clusters_type2_oracle, oracles_type1, oracles_type2, 
            filter_level, all_result, dict_all_clones):
 

    for num_lines in range(2,6):
        for unknown in (True, False):
            print(num_lines, unknown)
            new_clusters_type1, new_predictions_type1, new_clusters_oracle_type1, new_oracles_type1=filter_classes(clusters_type1, predictions_type1, clusters_type1_oracle, oracles_type1, unknown, num_lines)

            new_clusters_type2, new_predictions_type2, new_clusters_oracle_type2, new_oracles_type2=filter_classes(clusters_type2, predictions_type2, clusters_type2_oracle, oracles_type2, unknown, num_lines)

            # print("SIZES")

            # print("{} CLUSTER TYPE 1".format(len(new_clusters_type1.keys())))
            # print("{} PREDICTION TYPE 1".format(len(new_predictions_type1.keys())))
            # print("{} CLUSTER ORACLE TYPE 1".format(len(new_clusters_oracle_type1.keys())))
            # print("{} ORACLE TYPE 1".format(len(new_oracles_type1.keys())))

            # print("{} CLUSTER TYPE 2".format(len(new_clusters_type2.keys())))
            # print("{} PREDICTION TYPE 2".format(len(new_predictions_type2.keys())))
            # print("{} CLUSTER ORACLE TYPE 2".format(len(new_clusters_oracle_type2.keys())))
            # print("{} ORACLE TYPE 2".format(len(new_oracles_type2.keys())))

            # type1
            result_type1, type1_cl=create_curr_row(new_clusters_type1, new_predictions_type1, 
                new_oracles_type1, filter_level)
            # type2
            result_type2, type2_cl=create_curr_row(new_clusters_type2, new_predictions_type2, 
                new_oracles_type2, filter_level)

            # print("RESULT 1")
            # print(result_type1)

            # print("RESULT 2")
            # print(result_type2)


            result_global=merge_type1_type2(result_type1, result_type2)

            key="{}_{}".format(num_lines, unknown)
            if key not in all_result:
                all_result[key]=list()


            all_result[key].append(result_global)

            dict_all_clones[key+"_t1"]=type1_cl
            dict_all_clones[key+"_t2"]=type2_cl

    return all_result, dict_all_clones

def create_curr_row(clusters_typeX, predictions_typeX, oracles_typeX, filter_level):
    
    filtering_level=filter_level
    num_predictions=len(predictions_typeX.keys())
    perfect_prediction=0
    typeX_clones=list() # prediction that are in a cluster of typeX
    typeX_clones_pp=list()
    typeX_clones_wp=list()
    typeX_clones_oracle=list()
    typeX_pp_only_pretrain=list()
    typeX_pp_only_finetuning=list()
    typeX_pp_both=list()
    typeX_wp_only_pretrain=list()
    typeX_wp_only_finetuning=list()
    typeX_wp_both=list()

    for k in predictions_typeX.keys():
        if predictions_typeX[k].is_perfect:
            perfect_prediction+=1
        if len(predictions_typeX[k].clusters)>0:
            typeX_clones.append(k)
            # count the typeX clones for perfect prediction and wrong prediction
            if predictions_typeX[k].is_perfect:
                typeX_clones_pp.append(k)
            else:
                typeX_clones_wp.append(k)

            # check if a pretraining method is in at least one cluster
            clusters=predictions_typeX[k].clusters
            has_pretraining=False
            has_finetuning=False

            for cluster in clusters:
                cluster_curr=clusters_typeX[cluster]
                if len(cluster_curr.pretraining)>0:
                    has_pretraining=True
                    break
            for cluster in clusters:
                cluster_curr=clusters_typeX[cluster]
                if len(cluster_curr.training)>0:
                    has_finetuning=True
                    break

            if has_pretraining and not has_finetuning:
                if predictions_typeX[k].is_perfect:
                    typeX_pp_only_pretrain.append(k)
                else:
                    typeX_wp_only_pretrain.append(k)

            if has_finetuning and not has_pretraining:
                if predictions_typeX[k].is_perfect:
                    typeX_pp_only_finetuning.append(k)
                else:
                    typeX_wp_only_finetuning.append(k)

            if has_finetuning and has_pretraining:
                if predictions_typeX[k].is_perfect:
                    typeX_pp_both.append(k)
                else:
                    typeX_wp_both.append(k)


        # check if the oracle is contained into a cluster
        if len(oracles_typeX[k].clusters) >0:
            typeX_clones_oracle.append(k)

    result=list() # values to insert on the row
    result.append(filter_level) # dataset (e.g. 0807)
    result.append(perfect_prediction) # perfect prediction
    result.append(num_predictions) # total number of predictions
    result.append(len(typeX_clones)) # number of predictions that belong to at least one cluster
    result.append(len(typeX_clones_oracle)) # number of clones that have a clone also for oracle (among the once having it for the prediction)
    result.append(len(typeX_clones_pp)) # number of pp that belong to at least one cluster
    result.append(len(typeX_pp_only_pretrain)) # among the pp belonging to at least 1 typeX clone how many have only elements from pretrain method in the clusters
    result.append(len(typeX_pp_only_finetuning)) # among the pp belonging to at least 1 typeX clone how many have only elements finetuning method in the clusters    
    result.append(len(typeX_pp_both)) # among the pp belonging to at least 1 typeX clone how many have elements both from pretraining and finetuning in the clusters    
    result.append(len(typeX_clones_wp)) # number of wp that belong to at least one cluster
    result.append(len(typeX_wp_only_pretrain)) # among the wp belonging to at least 1 typeX clone how many have only elements from pretrain method in the clusters
    result.append(len(typeX_wp_only_finetuning)) # among the wp belonging to at least 1 typeX clone how many have only elements finetuning method in the clusters    
    result.append(len(typeX_wp_both)) # among the wp belonging to at least 1 typeX clone how many have elements both from pretraining and finetuning in the clusters    



    # # of clusters only pretrain + # of cluster only finetuning 
    # + # of cluster pretrain and finetuning = # of clusters
    # this must be true for pp and wp

    if len(typeX_pp_only_pretrain) + len(typeX_pp_only_finetuning) + len(typeX_pp_both) != len(typeX_clones_pp):
        print("ERROR: NUMBER OF CLUSTERS ARE NOT OK")

    if len(typeX_wp_only_pretrain) + len(typeX_wp_only_finetuning) + len(typeX_wp_both) != len(typeX_clones_wp):
        print("ERROR: NUMBER OF CLUSTERS ARE NOT OK")

    # print(len(typeX_clones_pp))
    # print(len(typeX_pp_pretrain))
    # print(len(typeX_pp_finetuning))
    # print(len(check_pretrain_and_finetuning_pp))
    # print("_____")

    # print(len(typeX_clones_wp))
    # print(len(typeX_wp_pretrain))
    # print(len(typeX_wp_finetuning))
    # print(len(check_pretrain_and_finetuning_wp))
    # print("_____")

    # sys.exit(0)

    return result, typeX_clones


def read_file(path):
    '''
    read a file in path location
    '''
    file=open(path, "r")
    lines=file.readlines()
    file.close()
    result=[l.replace("\n", "") for l in lines]
    return result

def process_files(predictions_dict, clusters_dict, excel_files, excel_folder, processing_prediction):
    for excel_file in excel_files:

        # Clone ID  Internal ID Num Lines   File    Start Line  End Line    Cloned Part 
        # Code checked for duplicates Correct prediction  Test method Method method id    Masked method id

        print("OPENING {}".format(os.path.join(excel_folder, excel_file)))
        wb_obj = openpyxl.load_workbook(os.path.join(excel_folder, excel_file))
        print("OPENED")
        sheet_obj = wb_obj["Clones"]

        max_row = sheet_obj.max_row #number of rows in the file

        for row in range(2, max_row+1):
            cluster_id=sheet_obj.cell(row, 1).value # id of the cluster (from 0 to the number of clusters found by Simian)
            id_file=sheet_obj.cell(row, 2).value

            if id_file is None:
                print("Skip one row")
                continue

            num_lines=int(sheet_obj.cell(row, 3).value)
            file_name=sheet_obj.cell(row, 4).value

            if cluster_id not in clusters_dict.keys():
                curr_class=Cluster(cluster_id, num_lines)
                clusters_dict[cluster_id]=curr_class

            # add the file to the current cluster
            clusters_dict[cluster_id].add_element(id_file, file_name, processing_prediction)

            # add to the prediction the current cluster in which it is contained
            if is_prediction(file_name, processing_prediction):
                prediction_id=int(id_file)
                # predictions[prediction_id].print()

                predictions_dict[prediction_id].add_cluster(cluster_id)

    return predictions_dict, clusters_dict


def create_summary(args):

    input_folder=args.input_folder
    oracle_folder=args.oracle
    
    '''
    read all the files that contain general information (i.e. not related to a specific dataset==model)
    '''

    # read cyclomatic complexity
    CC=read_file(args.CC)
    print("read {} lines for Cyclomatic Complexity".format(len(CC)))
    # read length test methods
    length_test_methods=read_file(args.length)
    print("read {} lines for length test methods".format(len(length_test_methods)))
    # read length test methods
    length_mask=read_file(args.mask_length)
    print("read {} lines for length masks".format(len(length_mask)))
    # read masked and method mapping
    masked_and_method=read_file(args.masked_and_method)
    print("read {} lines for masked and methods".format(len(masked_and_method)))
    # read number of masked lines
    masked_and_method=read_file(args.masked_and_method)
    print("read {} lines for masked and methods".format(len(masked_and_method)))
    # read unknown masked
    unknown_token=read_file(args.unknown_token)
    print("read {} lines for unknown token".format(len(unknown_token)))

    dict_unknown=dict()
    for k in unknown_token:
        dict_unknown[int(k)]=1

    dict_result=dict() # dict containing all the files we want to write

    dict_all_clones=dict() # dict containing each prediction that is a clone (just for checking them)

    for folder in folders:
        # processing predictions
        processing_prediction=True
        print("processing {} folder".format(folder))

        clusters_type1=dict()
        predictions_type1=dict()

        # reading perfect predictions
        perfect_predictions=read_file(os.path.join(args.perfect_predictions, folder, "perfect_predictions.txt"))
        print("read {} lines for perfect predictions".format(len(perfect_predictions)))
        # reading the confidence of the model
        confidence=read_file(os.path.join(args.confidence, folder, "scores.txt"))
        print("read {} lines for confidence".format(len(confidence)))

        # reading prediction of the model
        prediction_path=os.path.join(args.prediction_and_target, folder, "predictions/file_prediction_0.java")
        prediction_map_path=os.path.join(args.prediction_and_target, folder, "predictions/file_prediction_map.txt")

        model_predictions=read_file(prediction_path)
        print("read {} lines for model predictions".format(len(model_predictions)))
        model_predictions_map=read_file(prediction_map_path)
        print("read {} lines for model predictions mapping".format(len(model_predictions_map)))

        model_predictions_map_dict=dict()
        for m in model_predictions_map:
            parts=m.split("_")
            model_predictions_map_dict[int(parts[1])]=(int(parts[2]), int(parts[3]))

        # reading targets

        target_path=os.path.join(args.prediction_and_target, folder, "targets/file_target_0.java")
        target_map_path=os.path.join(args.prediction_and_target, folder, "targets/file_target_map.txt")
        
        target=read_file(target_path)
        print("read {} lines for target".format(len(target)))
        target_map=read_file(target_map_path)
        print("read {} lines for model predictions mapping".format(len(target_map)))

        target_map_dict=dict()
        for m in target_map:
            parts=m.split("_")
            target_map_dict[int(parts[1])]=(int(parts[2]), int(parts[3]))

        # add the information about the prediction retrieved from all the files
        predictions_type1=add_predictions_metrics(predictions_type1, masked_and_method, perfect_predictions, CC, length_test_methods)
        predictions_type1=add_prediction_and_target(predictions_type1, model_predictions, model_predictions_map_dict, target, target_map_dict)
        predictions_type1=add_unknown_token_and_line_masked(predictions_type1, dict_unknown, length_mask)
        predictions_type1=add_confidence(predictions_type1, confidence)

        excel_files=os.listdir(os.path.join(input_folder, folder))
        excel_files=[f for f in excel_files if ".xlsx" in f]

        type1=[f for f in excel_files if "type1" in f]
        type2=[f for f in excel_files if "type2" in f]

        # processing type1 files
        predictions_type1, clusters_type1 = process_files(predictions_type1, clusters_type1, type1, os.path.join(input_folder, folder), processing_prediction)

        num_cluster_1=len(clusters_type1)

        clusters_type2=dict()
        predictions_type2=dict()

        predictions_type2=add_predictions_metrics(predictions_type2, masked_and_method, perfect_predictions, CC, length_test_methods)
        predictions_type2=add_prediction_and_target(predictions_type2, model_predictions, model_predictions_map_dict, target, target_map_dict)
        predictions_type2=add_unknown_token_and_line_masked(predictions_type2, dict_unknown, length_mask)
        predictions_type2=add_confidence(predictions_type2, confidence)

        # processing type2 files
        predictions_type2, clusters_type2 = process_files(predictions_type2, clusters_type2, type2, os.path.join(input_folder, folder), processing_prediction)
 
        num_cluster_2=len(clusters_type2)


        # processing oracles
        processing_prediction=False

        print("processing {} folder".format(folder))

        clusters_type1_oracle=dict()
        oracles_type1=dict()
        clusters_type2_oracle=dict()
        oracles_type2=dict()

        excel_files=os.listdir(os.path.join(oracle_folder, folder))
        excel_files=[f for f in excel_files if ".xlsx" in f]

        type1=[f for f in excel_files if "type1" in f]
        type2=[f for f in excel_files if "type2" in f]

        print("OK")
        for i in range(len(perfect_predictions)):
            oracle=Oracle(i)
            oracles_type1[i]=oracle

        for i in range(len(perfect_predictions)):
            oracle=Oracle(i)
            oracles_type2[i]=oracle

        # processing type1 files
        oracles_type1, clusters_type1_oracle = process_files(oracles_type1, clusters_type1_oracle, type1, os.path.join(oracle_folder, folder), processing_prediction)


        print("{} clusters type 1".format(len(clusters_type1.keys())))
        # print("PRINT READ CLUSTERS")
        # for k in clusters_type1.keys():
        #     print(k)
        #     print(clusters_type1[k].pretraining)
        #     print(clusters_type1[k].training)
        #     print(clusters_type1[k].predictions)

        print("{} predictions type 1".format(len(predictions_type1.keys())))

        print("{} clusters oracle type 1".format(len(clusters_type1_oracle.keys())))
        print("{} oracle type 1".format(len(oracles_type1.keys())))


        # processing type2 files
        oracles_type2, clusters_type2_oracle = process_files(oracles_type2, clusters_type2_oracle, type2, os.path.join(oracle_folder, folder), processing_prediction)


        print("{} clusters type 2".format(len(clusters_type2.keys())))
        print("{} predictions type 2".format(len(predictions_type2.keys())))
 
        print("{} clusters oracle type 2".format(len(clusters_type2_oracle.keys())))
        print("{} oracle type 2".format(len(oracles_type2.keys())))



        columns=["Filter Level", "Perfect Predictions", "Total Number of Prediction",
        "Type 1 Clones", "Type 1 Clones In Oracle", "Type 1 Clones PP", "Type 1 Clones PP Only Pretrain", "Type 1 Clones PP Only Finetuning",
        "Type 1 Clones PP Pretrain And Finetuning", "Type 1 Clones WP", "Type 1 Clones WP Only Pretrain", "Type 1 Clones WP Only Finetuning",
        "Type 1 Clones WP Pretrain And Finetuning", "Type 2 Clones", "Type 2 Clones In Oracle", "Type 2 Clones PP", "Type 2 Clones PP Only Pretrain", "Type 2 Clones PP Only Finetuning",
        "Type 2 Clones PP Pretrain And Finetuning", "Type 2 Clones WP", "Type 2 Clones WP Only Pretrain", "Type 2 Clones WP Only Finetuning",
        "Type 2 Clones WP Pretrain And Finetuning"]


        dict_result, dict_all_clones=create_all_excels(clusters_type1, clusters_type2, predictions_type1, predictions_type2,
            clusters_type1_oracle, clusters_type2_oracle, oracles_type1, oracles_type2, folder, dict_result, dict_all_clones)


    for k in dict_result.keys():
        print(k)
        for dataset in dict_result[k]:

            for x,y in zip(columns, dataset):
                print("{}: {}".format(x,y))

    write_excels(args.output_folder, columns, dict_result)

    for k in dict_all_clones.keys():
        write_list(os.path.join(args.output_folder, "{}.txt".format(k)), dict_all_clones[k])

if __name__ == "__main__":


    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--input_folder",
        type=str,
        help="folder containing 0302,0403,0605,0807 folders, each of them with the excel files generated in step 3"
    )

    parser.add_argument(
        "--oracle",
        type=str,
        help="folder containing 0302,0403,0605,0807 folders, each of them with the excel files generated in step 9"
    )
  
    parser.add_argument(
        "--CC",
        type=str,
        help="file containing the cyclomatic complexity"
    )

    parser.add_argument(
        "--length",
        type=str,
        help="file containing the length of the test methods"
    )

    parser.add_argument(
        "--mask_length",
        type=str,
        help="file containing the length of the masked parts(from 2 to 6 lines)"
    )

    parser.add_argument(
        "--masked_and_method",
        type=str,
        help="file containing the mapping between masked method id and method id"
    )

    parser.add_argument(
        "--perfect_predictions",
        type=str,
        help="folder containing for each prediction if it is perfect or not (here we need the same folder contained in input folder)"
    )

    parser.add_argument(
        "--prediction_and_target",
        type=str,
        help="folder containing the folders with prediction and targets (here we need the same folder contained in input folder)"
    )

    parser.add_argument(
        "--output_folder",
        type=str,
        help="output folder where you want to write the summary file (here we need the same folder contained in input folder)"
    )

    parser.add_argument(
        "--unknown_token",
        type=str,
        help="file containing all the prediction with unknown token"
    )

    parser.add_argument(
        "--confidence",
        type=str,
        help="folder containing the confidence on the prediction (here we need the same folder contained in input folder)"
    )

    args = parser.parse_args()

    create_summary(args)


