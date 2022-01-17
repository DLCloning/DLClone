import os

import sys
import argparse

import math

import openpyxl
from openpyxl.styles import Font, Alignment

from utilities_summary1.excel_API import *
from utilities_summary1.prediction_cluster_class import Prediction, Cluster
from utilities_summary1.prediction_utilities import *

# folders you want to process
# folders=["test"]
folders=["0302", "0403", "0605", "0807"]

def write_excel(output_folder, curr_folder,  predictions_type1, clusters_type1,  predictions_type2, clusters_type2, num_cluster_1, num_cluster_2):
    print(os.path.join(output_folder, curr_folder))
    os.makedirs(os.path.join(output_folder, curr_folder), exist_ok=True)

    wb=create_excel_file()
    wb.create_sheet('Summary')
    sheet = wb["Summary"]

    fill_cells(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1", "M1", "N1", "O1", "P1", "Q1", "R1", "S1", "T1"],
    ["Prediction ID", "Is Perfect Prediction", "CC Test Method", "CC No Masked", "Num Lines", "Num Masked Lines", "Has Unknown Token", "Confidence", 
    "Prediction", "Target", "Num Clusters Type 1", "Tot Cluster Type 1", "Average Size Clusters Type 1", "Tot Elements Pretraining Type 1", 
    "Tot Elements Finetuning Type 1", "Num Clusters Type 2", 
    "Tot Cluster Type 2", "Average Size Clusters Type 2", "Tot Elements Pretraining Type 1", "Tot Elements Finetuning Type 1" ])

    make_bold(sheet, ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1", "M1", "N1", "O1", "P1", "Q1", "R1", "S1", "T1"])

    sheet.column_dimensions["I"].width = 50
    sheet.column_dimensions["J"].width = 50

    for i, k in enumerate(predictions_type1.keys()):
        predictions_type1[k].write_excel_line(i+2, sheet, clusters_type1, num_cluster_1)
        distinct_clusters=list(set(predictions_type2[k].clusters))
        average_size_cluster=predictions_type2[k].compute_average_cluster_size(clusters_type2)

        tot_number_pretraining, tot_number_finetuning=predictions_type2[k].compute_size_pretraining_finetuning(clusters_type2)


        fill_cells(sheet, ["P{}".format(i+2)], [len(distinct_clusters)])
        fill_cells(sheet, ["Q{}".format(i+2)], [num_cluster_2])
        fill_cells(sheet, ["R{}".format(i+2)], [average_size_cluster])
        fill_cells(sheet, ["S{}".format(i+2)], [tot_number_pretraining])
        fill_cells(sheet, ["T{}".format(i+2)], [tot_number_finetuning])


    wb.save(os.path.join(output_folder, curr_folder, "{}.xlsx".format(curr_folder)))

def read_file(path):
    '''
    read a file in path location
    '''
    file=open(path, "r")
    lines=file.readlines()
    file.close()
    result=[l.replace("\n", "") for l in lines]
    return result

def create_summary(args):

    input_folder=args.input_folder
    
    '''
    read all the files that contain general information (i.e. not related to a specific dataset==model)
    '''

    # read cyclomatic complexity
    CC=read_file(args.CC)
    print("read {} lines for Cyclomatic Complexity".format(len(CC)))
    # read length test methods
    length_test_methods=read_file(args.length)
    print("read {} lines for length test methods".format(len(length_test_methods)))
    # read length test methods
    length_mask=read_file(args.mask_length)
    print("read {} lines for length masks".format(len(length_mask)))
    # read masked and method mapping
    masked_and_method=read_file(args.masked_and_method)
    print("read {} lines for masked and methods".format(len(masked_and_method)))
    # read number of masked lines
    masked_and_method=read_file(args.masked_and_method)
    print("read {} lines for masked and methods".format(len(masked_and_method)))
    # read unknown masked
    unknown_token=read_file(args.unknown_token)
    print("read {} lines for unknown token".format(len(unknown_token)))

    dict_unknown=dict()
    for k in unknown_token:
        dict_unknown[int(k)]=1

    for folder in folders:
        print("processing {} folder".format(folder))

        clusters=dict()
        predictions=dict()

        # reading perfect predictions
        perfect_predictions=read_file(os.path.join(args.perfect_predictions, folder, "perfect_predictions.txt"))
        print("read {} lines for perfect predictions".format(len(perfect_predictions)))
        # reading the confidence of the model
        confidence=read_file(os.path.join(args.confidence, folder, "scores.txt"))
        print("read {} lines for confidence".format(len(confidence)))

        # reading prediction of the model
        prediction_path=os.path.join(args.prediction_and_target, folder, "predictions/file_prediction_0.java")
        prediction_map_path=os.path.join(args.prediction_and_target, folder, "predictions/file_prediction_map.txt")

        model_predictions=read_file(prediction_path)
        print("read {} lines for model predictions".format(len(model_predictions)))
        model_predictions_map=read_file(prediction_map_path)
        print("read {} lines for model predictions mapping".format(len(model_predictions_map)))

        model_predictions_map_dict=dict()
        for m in model_predictions_map:
            parts=m.split("_")
            model_predictions_map_dict[int(parts[1])]=(int(parts[2]), int(parts[3]))

        # reading targets

        target_path=os.path.join(args.prediction_and_target, folder, "targets/file_target_0.java")
        target_map_path=os.path.join(args.prediction_and_target, folder, "targets/file_target_map.txt")
        
        target=read_file(target_path)
        print("read {} lines for target".format(len(target)))
        target_map=read_file(target_map_path)
        print("read {} lines for model predictions mapping".format(len(target_map)))

        target_map_dict=dict()
        for m in target_map:
            parts=m.split("_")
            target_map_dict[int(parts[1])]=(int(parts[2]), int(parts[3]))

        # add the information about the prediction retrieved from all the files
        predictions=add_predictions_metrics(predictions, masked_and_method, perfect_predictions, CC, length_test_methods)
        predictions=add_prediction_and_target(predictions, model_predictions, model_predictions_map_dict, target, target_map_dict)
        prediction=add_unknown_token_and_line_masked(predictions, dict_unknown, length_mask)
        prediction=add_confidence(predictions, confidence)

        excel_files=os.listdir(os.path.join(input_folder, folder))
        excel_files=[f for f in excel_files if ".xlsx" in f]

        type1=[f for f in excel_files if "type1" in f]
        type2=[f for f in excel_files if "type2" in f]

        for excel_file in type1:

            # Clone ID  Internal ID Num Lines   File    Start Line  End Line    Cloned Part 
            # Code checked for duplicates Correct prediction  Test method Method method id    Masked method id

            print("OPENING {}".format(excel_file))
            wb_obj = openpyxl.load_workbook(os.path.join(input_folder, folder, excel_file))
            print("OPENED")
            sheet_obj = wb_obj["Clones"]

            max_row = sheet_obj.max_row #number of rows in the file

            for row in range(2, max_row+1):
                cluster_id=sheet_obj.cell(row, 1).value # id of the cluster (from 0 to the number of clusters found by Simian)
                id_file=sheet_obj.cell(row, 2).value

                if id_file is None:
                    print("Skip one row")
                    continue

                num_lines=int(sheet_obj.cell(row, 3).value)
                file_name=sheet_obj.cell(row, 4).value

                if cluster_id not in clusters.keys():
                    curr_class=Cluster(cluster_id, num_lines)
                    clusters[cluster_id]=curr_class

                # add the file to the current cluster
                clusters[cluster_id].add_element(id_file, file_name)

                # add to the prediction the current cluster in which it is contained
                if is_prediction(file_name):
                    prediction_id=int(id_file)
                    # predictions[prediction_id].print()

                    predictions[prediction_id].add_cluster(cluster_id)

        clusters_type1=clusters
        predictions_type1=predictions
        num_cluster_1=len(clusters_type1)

        clusters=dict()
        predictions=dict()

        predictions=add_predictions_metrics(predictions, masked_and_method, perfect_predictions, CC, length_test_methods)
        predictions=add_prediction_and_target(predictions, model_predictions, model_predictions_map_dict, target, target_map_dict)
        prediction=add_unknown_token_and_line_masked(predictions, dict_unknown, length_mask)
        prediction=add_confidence(predictions, confidence)
 
        for excel_file in type2:

            # Clone ID  Internal ID Num Lines   File    Start Line  End Line    Cloned Part 
            # Code checked for duplicates Correct prediction  Test method Method method id    Masked method id

            print("OPENING {}".format(excel_file))
            wb_obj = openpyxl.load_workbook(os.path.join(input_folder, folder, excel_file))
            print("OPENED")
            sheet_obj = wb_obj["Clones"]

            max_row = sheet_obj.max_row

            for row in range(2, max_row+1):
                cluster_id=sheet_obj.cell(row, 1).value
                id_file=sheet_obj.cell(row, 2).value

                if id_file is None:
                    print("Skip one row")
                    continue
                num_lines=int(sheet_obj.cell(row, 3).value)
                file_name=sheet_obj.cell(row, 4).value

                if cluster_id not in clusters.keys():
                    curr_class=Cluster(cluster_id, num_lines)
                    clusters[cluster_id]=curr_class

                clusters[cluster_id].add_element(id_file, file_name)

                if is_prediction(file_name):
                    prediction_id=int(id_file)
                    predictions[prediction_id].add_cluster(cluster_id)

            # for c in clusters.keys():
            #     clusters[c].print()

        num_cluster_2=len(clusters)

        write_excel(args.output_folder, folder, predictions_type1, clusters_type1, predictions, clusters, num_cluster_1, num_cluster_2)

if __name__ == "__main__":


    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--input_folder",
        type=str,
        help="folder containing 0302,0403,0605,0807 folders, each of them with the excel files generated in step 3"
    )
    
    parser.add_argument(
        "--CC",
        type=str,
        help="file containing the cyclomatic complexity"
    )

    parser.add_argument(
        "--length",
        type=str,
        help="file containing the length of the test methods"
    )

    parser.add_argument(
        "--mask_length",
        type=str,
        help="file containing the length of the masked parts(from 2 to 6 lines)"
    )

    parser.add_argument(
        "--masked_and_method",
        type=str,
        help="file containing the mapping between masked method id and method id"
    )

    parser.add_argument(
        "--perfect_predictions",
        type=str,
        help="folder containing for each prediction if it is perfect or not (here we need the same folder contained in input folder)"
    )

    parser.add_argument(
        "--prediction_and_target",
        type=str,
        help="folder containing the folders with prediction and targets (here we need the same folder contained in input folder)"
    )

    parser.add_argument(
        "--output_folder",
        type=str,
        help="output folder where you want to write the summary file (here we need the same folder contained in input folder)"
    )

    parser.add_argument(
        "--unknown_token",
        type=str,
        help="file containing all the prediction with unknown token"
    )

    parser.add_argument(
        "--confidence",
        type=str,
        help="folder containing the confidence on the prediction (here we need the same folder contained in input folder)"
    )

    args = parser.parse_args()

    create_summary(args)


